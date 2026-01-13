import logging
import openpyxl as xl
import os

from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string

from x2gtfs.config import Configuration

from x2gtfs.models import Stop, Calendar, CalendarDate, Agency, Route, Trip, StopTime
from x2gtfs.iterator import iter_data_vertical, iter_data_horizontal

def _parse_datetime(date_str: str, date_format: str) -> datetime:
    dt: datetime = datetime.strptime(date_str, date_format)

    return dt

def load_stop_metadata() -> dict[str, Stop]:
    stop_result_list: dict[str, Stop] = {}
    
    wb: Workbook = xl.load_workbook(
        os.path.join(Configuration.config.metadata.stops.input_filename), 
        data_only=True
    )

    ws: Worksheet = wb.active

    stop_identification_col_idx: int = column_index_from_string(Configuration.config.metadata.stops.stop_identification_index)
    stop_id_col_idx: int = column_index_from_string(Configuration.config.metadata.stops.stop_id_index)
    stop_name_col_idx: int = column_index_from_string(Configuration.config.metadata.stops.stop_name_index)
    stop_lat_col_idx: int = column_index_from_string(Configuration.config.metadata.stops.stop_lat_index)
    stop_lon_col_idx: int = column_index_from_string(Configuration.config.metadata.stops.stop_lon_index)
    
    current_row: int = 2
    while True:
        stop_identification_cell = ws.cell(row=current_row, column=stop_identification_col_idx)
        if stop_identification_cell.value is None:
            break
        
        stop = Stop()
        stop.stop_id = str(ws.cell(row=current_row, column=stop_id_col_idx).value)
        stop.stop_name = ws.cell(row=current_row, column=stop_name_col_idx).value
        stop.stop_lat = float(ws.cell(row=current_row, column=stop_lat_col_idx).value)
        stop.stop_lon = float(ws.cell(row=current_row, column=stop_lon_col_idx).value)

        stop_result_list[stop_identification_cell.value] = stop

        current_row += 1

    return stop_result_list

def load_calendar_metadata() -> tuple[dict[str, Calendar], dict[str, list[CalendarDate]]]:
    calendar_result_list: dict[str, Calendar] = {}
    calendar_date_result_list: dict[str, list[CalendarDate]] = {}

    wb: Workbook = xl.load_workbook(
        os.path.join(Configuration.config.metadata.calendars.input_filename), 
        data_only=True
    )

    ws: Worksheet = wb.active

    service_identification_col_idx: int = column_index_from_string(Configuration.config.metadata.calendars.service_identification_index)
    monday_col_idx: int = column_index_from_string(Configuration.config.metadata.calendars.monday_index)
    tuesday_col_idx: int = column_index_from_string(Configuration.config.metadata.calendars.tuesday_index)
    wednesday_col_idx: int = column_index_from_string(Configuration.config.metadata.calendars.wednesday_index) 
    thursday_col_idx: int = column_index_from_string(Configuration.config.metadata.calendars.thursday_index)
    friday_col_idx: int = column_index_from_string(Configuration.config.metadata.calendars.friday_index)
    saturday_col_idx: int = column_index_from_string(Configuration.config.metadata.calendars.saturday_index)
    sunday_col_idx: int = column_index_from_string(Configuration.config.metadata.calendars.sunday_index)
    start_date_col_idx: int = column_index_from_string(Configuration.config.metadata.calendars.start_date_index)
    end_date_col_idx: int = column_index_from_string(Configuration.config.metadata.calendars.end_date_index)

    current_row: int = 2
    while True:
        calendar: Calendar = Calendar()
        service_identification_cell = ws.cell(row=current_row, column=service_identification_col_idx)
        if service_identification_cell.value is None:
            break

        calendar.service_id = Configuration.config.defaults.service_id.format(service_id=len(calendar_result_list) + 1)
        calendar.monday = Configuration.config.mappings.calendar_day_types.dict().get(ws.cell(row=current_row, column=monday_col_idx).value, 0)
        calendar.tuesday = Configuration.config.mappings.calendar_day_types.dict().get(ws.cell(row=current_row, column=tuesday_col_idx).value, 0)
        calendar.wednesday = Configuration.config.mappings.calendar_day_types.dict().get(ws.cell(row=current_row, column=wednesday_col_idx).value, 0)
        calendar.thursday = Configuration.config.mappings.calendar_day_types.dict().get(ws.cell(row=current_row, column=thursday_col_idx).value, 0)
        calendar.friday = Configuration.config.mappings.calendar_day_types.dict().get(ws.cell(row=current_row, column=friday_col_idx).value, 0)
        calendar.saturday = Configuration.config.mappings.calendar_day_types.dict().get(ws.cell(row=current_row, column=saturday_col_idx).value, 0)
        calendar.sunday = Configuration.config.mappings.calendar_day_types.dict().get(ws.cell(row=current_row, column=sunday_col_idx).value, 0)
        
        start_date_value: str|datetime = ws.cell(row=current_row, column=start_date_col_idx).value
        if isinstance(start_date_value, datetime):
            calendar.start_date = start_date_value.strftime('%Y%m%d')
        else:
            calendar.start_date = _parse_datetime(  
                start_date_value,
                Configuration.config.metadata.calendars.date_format
            ).strftime('%Y%m%d')

        end_date_value: str|datetime = ws.cell(row=current_row, column=end_date_col_idx).value
        if isinstance(end_date_value, datetime):
            calendar.end_date = end_date_value.strftime('%Y%m%d')
        else:
            calendar.end_date = _parse_datetime(  
                end_date_value,
                Configuration.config.metadata.calendars.date_format
            ).strftime('%Y%m%d')

        calendar_result_list[service_identification_cell.value] = calendar

        current_row += 1

    wb: Workbook = xl.load_workbook(
        os.path.join(Configuration.config.metadata.calendar_exceptions.input_filename), 
        data_only=True
    )

    ws: Worksheet = wb.active

    service_identification_col_idx: int = column_index_from_string(Configuration.config.metadata.calendar_exceptions.service_identification_index)
    date_col_idx: int = column_index_from_string(Configuration.config.metadata.calendar_exceptions.date_index)
    exception_type_col_idx: int = column_index_from_string(Configuration.config.metadata.calendar_exceptions.exception_type_index)

    current_row: int = 2
    while True:
        calendar_date: CalendarDate = CalendarDate()
        service_identification_cell = ws.cell(row=current_row, column=service_identification_col_idx)
        if service_identification_cell.value is None:
            break

        if service_identification_cell.value not in calendar_result_list:
            continue

        calendar_date.service_id = calendar_result_list[service_identification_cell.value].service_id
        
        date_value: str|datetime = ws.cell(row=current_row, column=date_col_idx).value
        if isinstance(date_value, datetime):
            calendar_date.date = date_value.strftime('%Y%m%d')
        else:
            calendar_date.date = _parse_datetime(  
                date_value,
                Configuration.config.metadata.calendar_exceptions.date_format
            ).strftime('%Y%m%d')

        calendar_date.exception_type = Configuration.config.mappings.calendar_exception_types.dict().get(
            ws.cell(row=current_row, column=exception_type_col_idx).value,
            1
        )

        if service_identification_cell.value not in calendar_date_result_list:
            calendar_date_result_list[service_identification_cell.value] = []

        calendar_date_result_list[service_identification_cell.value].append(calendar_date)

        current_row += 1

    return calendar_result_list, calendar_date_result_list

def load_agency_and_route_metadata() -> tuple[dict[str, Agency], dict[str, Route]]:
    agency_result_list: dict[str, Agency] = {}
    route_result_list: dict[str, Route] = {}

    wb: Workbook = xl.load_workbook(
        os.path.join(Configuration.config.metadata.routes.input_filename), 
        data_only=True
    )

    ws: Worksheet = wb.active

    route_identification_col_idx: int = column_index_from_string(Configuration.config.metadata.routes.route_identification_index)
    route_id_col_idx: int = column_index_from_string(Configuration.config.metadata.routes.route_id_index)
    route_short_name_col_idx: int = column_index_from_string(Configuration.config.metadata.routes.route_short_name_index)
    route_long_name_col_idx: int = column_index_from_string(Configuration.config.metadata.routes.route_long_name_index)
    route_type_col_idx: int = column_index_from_string(Configuration.config.metadata.routes.route_type_index)
    route_color_col_idx: int = column_index_from_string(Configuration.config.metadata.routes.route_color_index)
    route_text_color_col_idx: int = column_index_from_string(Configuration.config.metadata.routes.route_text_color_index)
    agency_name_col_idx: int = column_index_from_string(Configuration.config.metadata.routes.agency_name_index)
    agency_url_col_idx: int = column_index_from_string(Configuration.config.metadata.routes.agency_url_index)

    current_row: int = 2
    while True:
        route_identification_cell = ws.cell(row=current_row, column=route_identification_col_idx)
        if route_identification_cell.value is None:
            break
        
        route = Route()
        route.route_id = str(ws.cell(row=current_row, column=route_id_col_idx).value)
        route.route_short_name = ws.cell(row=current_row, column=route_short_name_col_idx).value
        route.route_long_name = ws.cell(row=current_row, column=route_long_name_col_idx).value
        
        if ws.cell(row=current_row, column=route_type_col_idx).value in Configuration.config.mappings.route_type.dict().keys():
            route.route_type = Configuration.config.mappings.route_type.dict()[ws.cell(row=current_row, column=route_type_col_idx).value]
        else:
            route.route_type = 3

        route.route_color = ws.cell(row=current_row, column=route_color_col_idx).value
        route.route_text_color = ws.cell(row=current_row, column=route_text_color_col_idx).value

        route_result_list[route_identification_cell.value] = route

        if ws.cell(row=current_row, column=agency_name_col_idx).value not in agency_result_list.keys():
            agency = Agency()
            agency.agency_id = Configuration.config.defaults.agency_id.format(agency_id=len(agency_result_list) + 1)
            agency.agency_name = ws.cell(row=current_row, column=agency_name_col_idx).value
            agency.agency_url = ws.cell(row=current_row, column=agency_url_col_idx).value
            agency.agency_timezone = Configuration.config.defaults.agency_timezone

            route.agency_id = agency.agency_id

            agency_result_list[route_identification_cell.value] = agency
        else:
            route.agency_id = agency_result_list[ws.cell(row=current_row, column=agency_name_col_idx).value].agency_id

        current_row += 1

    return agency_result_list, route_result_list

def process_timetable_files(
        stop_meta_list: dict[str, Stop],
        calendar_meta_list: dict[str, Calendar], 
        calendar_exception_meta_list: dict[str, list[CalendarDate]], 
        agency_meta_list: dict[str, Agency],
        route_meta_list: dict[str, Route]) -> tuple[list[Trip], list[StopTime]]:
    
    trip_result_list: list[Trip] = []
    stop_time_result_list: list[StopTime] = []

    for input_filename in os.listdir(Configuration.config.timetables.input_directory):
        if input_filename.endswith('.xlsx'):
            logging.info(f"Processing file: {input_filename}")
            wb: Workbook = xl.load_workbook(
                os.path.join(Configuration.config.timetables.input_directory, input_filename), 
                data_only=True
            )

            ws: Worksheet = wb.active

            current_trip_idx: int = -1
            current_stop_idx: str|None = None

            current_trip: Trip|None = None
            current_stop_time: StopTime|None = None

            if Configuration.config.timetables.layout_type == 'vertical':
                for cell in iter_data_vertical(ws, ws[Configuration.config.timetables.data_start_area]):
                    if cell.value == Configuration.config.timetables.run_through_char:
                        continue
                    
                    if not cell.column == current_trip_idx:
                        if current_trip is not None:
                            trip_result_list.append(current_trip)
                        
                        current_trip = Trip()
                        current_trip.trip_id = Configuration.config.defaults.trip_id.format(trip_id=len(trip_result_list) + 1).zfill(6)
                        
                        route_idx: str = ws.cell(Configuration.config.timetables.route_identification_index, cell.column).value
                        if route_idx in route_meta_list:
                            current_trip.route_id = route_meta_list[route_idx].route_id
                        else:
                            logging.warning(f"Route identification '{route_idx}' not found in route metadata. Using route identification as route_id fallback.")
                            current_trip.route_id = route_idx
                        
                        service_idx: str = ws.cell(Configuration.config.timetables.service_identification_index, cell.column).value
                        if service_idx in calendar_meta_list:
                            current_trip.service_id = calendar_meta_list[service_idx].service_id
                        elif service_idx in calendar_exception_meta_list:
                            current_trip.service_id = calendar_exception_meta_list[service_idx][0].service_id
                        else:
                            logging.warning(f"Service identification '{service_idx}' not found in calendar metadata. Using service identification as service_id fallback.")
                            current_trip.service_id = service_idx

                        shape_idx: str = ws.cell(Configuration.config.timetables.shape_identification_index, cell.column).value

                        current_trip.trip_short_name = ws.cell(Configuration.config.timetables.trip_short_name_index, cell.column).value
                        current_trip.trip_headsign = ws.cell(Configuration.config.timetables.trip_headsign_index, cell.column).value

                        current_trip_idx = cell.column
                        
                    stop_idx: str = ws.cell(cell.row, column_index_from_string(Configuration.config.timetables.stop_identification_index)).value
                    if not stop_idx == current_stop_idx:
                        current_stop_time = StopTime()
                        current_stop_time.trip_id = current_trip.trip_id

                        if stop_idx in stop_meta_list:
                            current_stop_time.stop_id = stop_meta_list[stop_idx].stop_id
                        else:
                            logging.warning(f"Stop identification '{stop_idx}' not found in stop metadata. Using stop identification as stop_id fallback.")
                            current_stop_time.stop_id = stop_idx

                        current_stop_time.stop_sequence = len([st for st in stop_time_result_list if st.trip_id == current_trip.trip_id]) + 1
                        current_stop_time.arrival_time = cell.value
                        current_stop_time.departure_time = cell.value

                        stop_time_result_list.append(current_stop_time)
                        current_stop_idx = stop_idx
                    else:
                        if stop_idx in stop_meta_list:
                            resolved_stop_id: str = stop_meta_list[stop_idx].stop_id
                        else:
                            resolved_stop_id: str = stop_idx

                        last_stop_time: StopTime = next(st for st in reversed(stop_time_result_list) if st.trip_id == current_trip.trip_id and st.stop_id == resolved_stop_id)
                        last_stop_time.departure_time = cell.value

                if current_trip is not None:
                    trip_result_list.append(current_trip)

            else:
                raise NotImplementedError("Only 'vertical' layout type is currently implemented.")

    return trip_result_list, stop_time_result_list