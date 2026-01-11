from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from typing import Generator

def iter_data_vertical(ws: Worksheet, start_cell: Cell) -> Generator[Cell, None, None]:
    """
    Iterate vertically over columns starting from start_cell.
    
    - Each column is traversed from start_cell.row downwards
    - Row iteration stops at the first empty cell after at least one value has been found
    - Then moves to the next column
    - Iteration stops only when an entire column from start_cell.row is empty
    
    Args:
        ws: openpyxl Worksheet object
        start_cell: The starting cell for the iteration

    Yields:
        Cell objects in the order of iteration
    """
    start_row: int = start_cell.row
    start_col: int = start_cell.column
    col: int = start_col
    max_row: int = ws.max_row

    while True:
        # Check if the entire column from start_row is empty
        col_empty: bool = True
        for r in range(start_row, max_row + 1):
            if ws.cell(row=r, column=col).value not in (None, ""):
                col_empty = False
                break

        if col_empty:
            # Entire column is empty → stop iteration
            break

        # Iterate rows in the current column
        row: int = start_row
        found_value_in_column: bool = False
        while row <= max_row:
            cell: Cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ""):
                found_value_in_column = True
                yield cell
            elif found_value_in_column:
                # After the first value, an empty cell → move to next column
                break
            row += 1

        col += 1

def iter_data_horizontal(ws: Worksheet, start_cell: Cell) -> Generator[Cell, None, None]:
    """
    Iterate horizontally over rows starting from start_cell.
    
    - Each row is traversed from start_cell.column to the right
    - Column iteration stops at the first empty cell after at least one value has been found
    - Then moves to the next row
    - Iteration stops only when an entire row from start_cell.column is empty
    
    Args:
        ws: openpyxl Worksheet object
        start_cell: The starting cell for the iteration

    Yields:
        Cell objects in the order of iteration
    """
    start_row: int = start_cell.row
    start_col: int = start_cell.column
    row: int = start_row
    max_col: int = ws.max_column

    while True:
        # Check if the entire row from start_col is empty
        row_empty: bool = True
        for c in range(start_col, max_col + 1):
            if ws.cell(row=row, column=c).value not in (None, ""):
                row_empty = False
                break

        if row_empty:
            # Entire row is empty → stop iteration
            break

        # Iterate columns in the current row
        col: int = start_col
        found_value_in_row: bool = False
        while col <= max_col:
            cell: Cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ""):
                found_value_in_row = True
                yield cell
            elif found_value_in_row:
                # After the first value, an empty cell → move to next row
                break
            col += 1

        row += 1